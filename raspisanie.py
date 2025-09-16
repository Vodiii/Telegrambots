import openpyxl as xl

def get_raspisanie(day=1) -> dict:
    wb = xl.load_workbook('files/lesson_schedule.xlsx')
    '''Получает расписание на день недели с номером day.
    Понеделник = 1, Вторник = 2 и т. д.
    
    Возвращает список с расписанием'''
    if type(day) != type(1): raise TypeError
    if day >= 8 or day <= 0: raise ValueError
    res = []
    for i in range(2, 8):
        t, v = wb["лист 1"].cell(row=i, column=1).value, wb["лист 1"].cell(row=i, column=day + 1).value
        if v:
            res.append([t, v]) 
        else:
            continue
    return res


if __name__ == "__main__":
    print(help(get_raspisanie))
    for i in get_raspisanie("g"):
        print(i)